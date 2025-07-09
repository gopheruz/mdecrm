import mimetypes
import fnmatch
import os
import re
from datetime import datetime, timedelta
from io import BytesIO
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch, Q, Count, F, Case, When, IntegerField
from django.db.models.functions import ExtractHour, TruncDay
from django.http import (
    JsonResponse, Http404, FileResponse, HttpResponse, 
    HttpResponseForbidden, HttpResponseServerError
)
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from med_app.forms import *
from med_app.models import *
from django.db.models import Q
import fnmatch
from django.contrib.auth.views import redirect_to_login

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from .models import MedCard, City, District
from .forms import MedCardForm,VisitEditForm

@login_required
def serve_recording_view(request, call_id):
    call_instance = get_object_or_404(Call, id=call_id)
    path_pattern = call_instance.recording_path

    if not path_pattern:
        messages.error(request, f"Для звонка #{call_instance.id} не указан путь к записи.")
        raise Http404("Путь-шаблон к записи для этого звонка не указан.")

    dir_name = os.path.dirname(path_pattern)
    file_pattern = os.path.basename(path_pattern)

    actual_file_path = None
    found_filename = None

    if not os.path.isdir(dir_name):
        messages.error(request, f"Директория для записей '{dir_name}' не найдена на сервере.")
        raise Http404(f"Директория для записей не найдена для звонка #{call_instance.id}.")

    try:
        possible_files = []
        for filename_on_server in os.listdir(dir_name):
            if fnmatch.fnmatch(filename_on_server, file_pattern):
                possible_files.append(filename_on_server)

        if not possible_files:
            messages.warning(request,
                             f"Файл записи не найден по шаблону '{path_pattern}' для звонка #{call_instance.id}.")
            raise Http404(f"Файл записи не найден для звонка #{call_instance.id}.")

        if len(possible_files) == 1:
            found_filename = possible_files[0]
            actual_file_path = os.path.join(dir_name, found_filename)
        else:
            messages.info(request,
                          f"Найдено несколько файлов ({len(possible_files)}) для звонка #{call_instance.id}. Использован первый: {possible_files[0]}")
            found_filename = sorted(possible_files)[0]
            actual_file_path = os.path.join(dir_name, found_filename)

    except OSError as e:
        messages.error(request, f"Ошибка доступа к директории записей для звонка #{call_instance.id}: {e}")
        raise Http404(f"Ошибка сервера при поиске файла записи для звонка #{call_instance.id}.")

    if not actual_file_path or not os.path.isfile(actual_file_path):
        messages.error(request,
                       f"Файл записи '{actual_file_path or found_filename}' не является файлом или не существует для звонка #{call_instance.id}.")
        raise Http404(f"Файл записи не найден или не доступен для звонка #{call_instance.id}.")

    content_type, encoding = mimetypes.guess_type(actual_file_path)
    if content_type is None:
        content_type = 'application/octet-stream'

    try:
        response = FileResponse(open(actual_file_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'inline; filename="{os.path.basename(actual_file_path)}"'
        return response
    except IOError:
        messages.error(request,
                       f"Ошибка чтения файла записи '{actual_file_path}' на сервере для звонка #{call_instance.id}.")
        raise Http404(f"Ошибка чтения файла записи для звонка #{call_instance.id}.")
    except Exception as e:
        messages.error(request, f"Неизвестная ошибка при доступе к файлу записи для звонка #{call_instance.id}: {e}")
        raise Http404(f"Неизвестная ошибка сервера при доступе к файлу записи для звонка #{call_instance.id}.")


@login_required
def get_call_questions_ajax(request):
    call_id = request.GET.get('call_id')
    if not call_id:
        return JsonResponse({'error': 'Missing call_id'}, status=400)

    try:
        call = Call.objects.get(id=call_id)
        call_questions = call.call_questions.select_related('department').prefetch_related('questions').all()
        data = []
        for cq in call_questions:
            questions_list = [{'id': q.id, 'question': q.question} for q in cq.questions.all()]
            if questions_list:
                data.append({
                    'department': cq.department.name if cq.department else 'Без отделения',
                    'questions': questions_list
                })
        return JsonResponse({'questions_data': data})

    except Call.DoesNotExist:
        return JsonResponse({'error': 'Call not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def custom_login_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path(), login_url='/user_app/login/')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

@custom_login_required
def get_questions(request):
    department_id = request.GET.get('department_id')
    questions = Question.objects.filter(department_id=department_id).values('id', 'question')
    return JsonResponse(list(questions), safe=False)


def index_view(request):
    context = {
        'title': 'Bosh sahifa',
        'is_staff_user': request.user.is_staff,
        'wav_files': [],
        'med_cards': [],
    }

    today = timezone.localtime(timezone.now()).date()
    context['today'] = today

    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
    else:
        selected_date = today

    context['selected_date'] = selected_date
    context['yesterday'] = selected_date - timedelta(days=1)
    context['day_before_yesterday'] = selected_date - timedelta(days=2)

    if request.user.is_authenticated:
        base_dir = '/mnt/cdr'
        local_dir = os.path.join(base_dir, f"{selected_date.year}", f"{selected_date.month:02d}", f"{selected_date.day:02d}")
        wav_files = []
        try:
            if os.path.exists(local_dir):
                for filename in os.listdir(local_dir):
                    if filename.endswith('.wav'):
                        file_path = os.path.join(local_dir, filename).replace('\\', '/')
                        try:
                            file_size = os.path.getsize(file_path)
                            if file_size > 44:  # Fayl hajmi > 44 bayt
                                wav_file = {
                                    'path': file_path,
                                    'filename': filename,
                                    'phone_number': '-',
                                    'ip_operator': '-',
                                    'call_time': '-',
                                    'file_size': file_size,
                                }

                                # out va q-1001 fayllar uchun moslashgan regex
                                pattern = r'(out|q-1001)-(\d+)(?:-(\d+))?-(\d{8})-(\d{6})-(\d+\.\d+)\.wav'
                                match = re.match(pattern, filename)

                                if match:
                                    file_prefix, first_num, second_num, date, time, unique_id = match.groups()

                                    if file_prefix == 'out':
                                        phone_number = first_num
                                        ip_operator = second_num or '-'
                                    else:  # q-1001
                                        phone_number = first_num
                                        ip_operator = "1001"

                                    wav_file['phone_number'] = f"+{phone_number}" if phone_number != '-' and not phone_number.startswith('+') else phone_number
                                    wav_file['ip_operator'] = ip_operator
                                    wav_file['call_time'] = f"{time[:2]}:{time[2:4]}:{time[4:6]}"

                                else:
                                    print(f"❌ Tahlil qilinmadi: {filename}")

                                wav_files.append(wav_file)
                            else:
                                print(f"⏭️ Fayl o‘tkazib yuborildi: {filename} (hajmi {file_size} bayt <= 44)")

                        except OSError as e:
                            print(f"DEBUG: Faylga kirishda xato: {filename}: {e}")  # Debug
            else:
                print(f"DEBUG: Direktoriya topilmadi: {local_dir}")  # Debug
            context['wav_files'] = sorted(wav_files, key=lambda x: x.get('call_time', ''), reverse=True)

        except Exception as e:
            print(f"DEBUG: Fayl tizimi xatosi: {e}")

        try:
            med_cards = MedCard.objects.filter(
                Q(call__created_at__date=selected_date) |
                Q(visits__visit_time__date=selected_date)
            ).distinct().select_related('city', 'district').order_by('last_name')
            if not med_cards.exists():
                med_cards = MedCard.objects.all().select_related('city', 'district')

            context['med_cards'] = med_cards
        except Exception as e:
            print(f"DEBUG: Tibbiy kartalarni olishda xato: {e}")  # Debug
    return render(request, 'med_app/index.html', context)

@custom_login_required
def stream_wav_file(request, wav_path):
    base_dir = os.path.join(os.getcwd(), '/mnt/cdr')
    wav_path = os.path.normpath(wav_path.lstrip('/')).replace('\\', '/')
    full_path = os.path.join(base_dir, wav_path)
    if not os.path.exists(full_path):
        raise Http404("Fayl topilmadi")

    return FileResponse(open(full_path, 'rb'), content_type='audio/wav')


def edit_med_card(request, pk):
    med_card = get_object_or_404(MedCard, pk=pk)
    
    if request.method == 'POST':
        form = MedCardForm(request.POST, instance=med_card)
        if form.is_valid():
            form.save()
            messages.success(request, 'Медицинская карта успешно обновлена!')
            return redirect('med_card_profile_url', id=pk)
    else:
        form = MedCardForm(instance=med_card)
    
    context = {
        'form': form,
        'med_card': med_card,
    }
    return render(request, 'med_app/edit_med_card.html', context)

@custom_login_required
def create_med_cart_get_view(request):
    form = CreateMedCartForm()
    context = {
        'title': 'Создание мед. карты !',
        'form': form
    }
    return render(request, 'med_app/create_med_cart.html', context)

@custom_login_required
def create_med_cart_post_view(request):
    form = CreateMedCartForm(request.POST)
    if form.is_valid():
        new_city_name = form.cleaned_data.get('new_city_name')
        new_district_name = form.cleaned_data.get('new_district_name')
        selected_city = form.cleaned_data.get('city')
        selected_district = form.cleaned_data.get('district')

        final_city = None
        final_district = None

        if new_city_name:
            city_obj, created = City.objects.get_or_create(
                name__iexact=new_city_name,
                defaults={'name': new_city_name}
            )
            final_city = city_obj
            if created:
                messages.info(request, f"Создан новый город: {final_city.name}")
        elif selected_city:
            final_city = selected_city

        if final_city:
            if new_district_name:
                district_obj, created = District.objects.get_or_create(
                    city=final_city,
                    name__iexact=new_district_name,
                    defaults={'name': new_district_name}
                )
                final_district = district_obj
                if created:
                    messages.info(request, f"Создан новый район: {final_district.name} в городе {final_city.name}")
            elif selected_district:
                if selected_district.city == final_city:
                    final_district = selected_district
                else:
                    messages.error(request, "Ошибка: Выбранный район не соответствует городу.")
                    return render(request, 'med_app/create_med_cart.html',
                                  {'form': form, 'title': 'Создать мед. карту'})

        if final_city and final_district:
            med_card = form.save(commit=False)
            med_card.city = final_city
            med_card.district = final_district
            med_card.save()
            messages.success(request, f"Медицинская карта для {med_card} успешно создана.")
            return redirect('med_card_profile_url', med_card.id)
        else:
            messages.error(request, "Не удалось определить город или район для сохранения медкарты.")
    else:
        messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    return render(request, 'med_app/create_med_cart.html', {
        'form': form,
        'title': 'Создать мед. карту'
    })

@custom_login_required
def med_card_profile_view(request, id):
    med_card = get_object_or_404(MedCard, id=id)
    base_dir = '/mnt/cdr'
    calls = []

    # Search all subdirectories for .wav files matching the med_card's phone number
    found_files = []
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.endswith('.wav'):
                # Regex for both out- and q-1001- formats
                out_pattern = r'^out-(\d+)-(\d+)-(\d{8})-(\d{6})-(.+)\.wav$'
                q_pattern = r'^q-1001-(\d+)-(\d{8})-(\d{6})-(.+)\.wav$'
                out_match = re.match(out_pattern, file)
                q_match = re.match(q_pattern, file)
                call_data = None

                if out_match and out_match.group(1) == str(med_card.phone_number):
                    phone_number = out_match.group(1)
                    operator = out_match.group(2)
                    date_str = out_match.group(3)
                    time_str = out_match.group(4)
                    dt_str = f"{date_str} {time_str}"
                    try:
                        created_at = datetime.strptime(dt_str, "%Y%m%d %H%M%S")
                    except Exception:
                        created_at = None
                    file_path = os.path.join(root, file)
                    if os.path.getsize(file_path) > 44:
                        relative_path = os.path.relpath(file_path, base_dir).replace('\\', '/')
                        call_data = {
                            'id': file,
                            'phone_number': phone_number,
                            'created_at': created_at,
                            'operator': operator,
                            'wav_path': relative_path,
                        }
                elif q_match and q_match.group(1) == str(med_card.phone_number):
                    phone_number = q_match.group(1)
                    operator = "1001"
                    date_str = q_match.group(2)
                    time_str = q_match.group(3)
                    dt_str = f"{date_str} {time_str}"
                    try:
                        created_at = datetime.strptime(dt_str, "%Y%m%d %H%M%S")
                    except Exception:
                        created_at = None
                    file_path = os.path.join(root, file)
                    if os.path.getsize(file_path) > 44:
                        relative_path = os.path.relpath(file_path, base_dir).replace('\\', '/')
                        call_data = {
                            'id': file,
                            'phone_number': phone_number,
                            'created_at': created_at,
                            'operator': operator,
                            'wav_path': relative_path,
                        }
                        # Agar aynan shu fayl bo'lsa, logga chiqaramiz
                        if file == "q-1001-994151224-20250707-095211-1751863931.502.wav":
                            print(f"✅ Fayl topildi: {file_path}")
                            found_files.append(file_path)
                if call_data:
                    calls.append(call_data)

    if not found_files:
        print("❌ q-1001-994151224-20250707-095211-1751863931.502.wav fayli topilmadi yoki hajmi <= 44 bayt")

    calls = sorted(calls, key=lambda x: x['created_at'] or datetime.min, reverse=True)
    visits = Visit.objects.filter(med_card=id)
    context = {
        'title': 'Мед. карта',
        'med_card': med_card,
        'calls': calls,
        'visits': visits
    }
    return render(request, 'med_app/med_card_profile.html', context)

@custom_login_required
def search_med_card_get_view(request):
    form = SearchMedCardForm()
    context = {
        'title': 'Поиск',
        'form': form
    }
    return render(request, 'med_app/search_med_card.html', context)

@custom_login_required
def search_med_card_post_view(request):
    form = SearchMedCardForm(request.POST or None)
    results = None
    if request.method == 'POST' and form.is_valid():
        first_name = form.cleaned_data['first_name']
        last_name = form.cleaned_data['last_name']
        surname = form.cleaned_data['surname']
        phone_number = form.cleaned_data['phone_number']
        birth_date = form.cleaned_data['birth_date']

        query = Q()
        if first_name:
            query &= Q(first_name__icontains=first_name)
        if last_name:
            query &= Q(last_name__icontains=last_name)
        if surname:
            query &= Q(surname__icontains=surname)
        if phone_number:
            query &= Q(phone_number__icontains=phone_number)
        if birth_date:
            query &= Q(birth_date=birth_date)

        if query:
            results = MedCard.objects.filter(query)

    context = {
        'form': form,
        'results': results,
        'title': 'Результаты поиска'
    }
    return render(request, 'med_app/search_med_card.html', context)

@custom_login_required
def create_call_get_or_post_view(request, med_card_id):
    med_card = get_object_or_404(MedCard, id=med_card_id)
    if request.method == 'POST':
        call_form = CallForm(request.POST)
        formset = CallQuestionFormSet(request.POST)
        if call_form.is_valid() and formset.is_valid():
            call = call_form.save(commit=False)
            call.operator = request.user
            call.med_card = med_card
            call.save()
            instances = formset.save(commit=False)
            for instance in instances:
                instance.call = call
                instance.save()
            formset.save_m2m()
            for obj in formset.deleted_objects:
                obj.delete()
            messages.success(request, 'Звонок создан успешно!')
            return redirect('index_url')
        else:
            messages.error(request, 'Исправьте ошибки в форме')
            context = {
                'title': 'Создание звонка (ошибка)',
                'call_form': call_form,
                'formset': formset,
                'med_card': med_card,
                'med_card_id': med_card_id
            }
            return render(request, 'med_app/create_call.html', context)
    else:
        call_form = CallForm()
        formset = CallQuestionFormSet(queryset=CallQuestion.objects.none())
        context = {
            'title': 'Создание звонка',
            'call_form': call_form,
            'formset': formset,
            'med_card': med_card,
            'med_card_id': med_card_id
        }
        return render(request, 'med_app/create_call.html', context)

@custom_login_required
def call_detail_view(request, call_id):
    call = Call.objects.get(id=call_id)
    context = {
        'title': 'История звонка',
        'call': call,
        'user': request.user,
        'call_questions': call.call_questions.all()
    }
    return render(request, 'med_app/call_detail.html', context)

@custom_login_required
def create_visit_get_view(request, med_card_id):
    form = VisitCreateForm()
    context = {
        'title': 'Создать посещение',
        'form': form,
        'med_card_id': med_card_id
    }
    return render(request, 'med_app/create_visit.html', context)

@custom_login_required
def create_visit_post_view(request, med_card_id):
    form = VisitCreateForm(data=request.POST)
    if form.is_valid():
        med_card = MedCard.objects.get(id=med_card_id)
        visit = form.save(commit=False)
        visit.med_card = med_card
        visit.operator = request.user
        visit.save()
        messages.info(request, "Визит успешно создан !")
        return redirect('med_card_profile_url', med_card_id)

def visit_edit(request, visit_id):
    visit = get_object_or_404(Visit, id=visit_id)
    if request.method == 'POST':
        form = VisitEditForm(request.POST, instance=visit)
        if form.is_valid():
            form.save()
            return redirect('med_card_profile_url', id=visit.med_card.id)
    else:
        form = VisitEditForm(instance=visit)
    return render(request, 'med_app/visit_edit.html', {'form': form, 'visit': visit})
@custom_login_required
def all_med_cards_view(request):
    med_cards = MedCard.objects.all()
    context = {
        'title': 'Все мед. карты',
        'med_cards': med_cards
    }
    return render(request, 'med_app/all_med_cards.html', context)

@custom_login_required
def reports_view(request):
    if request.method == 'GET':
        operator_form = OperatorReportForm()
        questions_form = QuestionsReportForm()
        visits_form = VisitsReportForm()
        context = {
            'operator_form': operator_form,
            'questions_form': questions_form,
            'visits_form': visits_form,
            'title': 'Отчеты',
        }
        return render(request, 'med_app/reports.html', context)

@custom_login_required
def operator_report_view(request):
    if request.method == 'POST':
        form = OperatorReportForm(request.POST)
        calls = None
        daily_calls = None
        title = "Отчет по операторам"
        if form.is_valid():
            operator = form.cleaned_data['operator']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            queryset = Call.objects.filter(operator=operator)
            if start_date and end_date:
                end_date_inclusive = datetime.combine(end_date, datetime.max.time())
                queryset = queryset.filter(created_at__range=(start_date, end_date_inclusive))
            if 'generate_report' in request.POST:
                calls = queryset.select_related(
                    'med_card__city',
                    'med_card__district',
                    'operator'
                ).prefetch_related(
                    Prefetch(
                        'call_questions',
                        queryset=CallQuestion.objects.select_related('department').prefetch_related('questions')
                    )
                ).order_by('-created_at')
                daily_calls = queryset.values('created_at__date').annotate(
                    count=Count('id')
                ).order_by('created_at__date')
                context = {
                    'title': title,
                    'operator_form': form,
                    'questions_form': QuestionsReportForm(),
                    'calls': calls,
                    'daily_calls': daily_calls,
                }
                return render(request, 'med_app/reports.html', context)
            elif 'export_excel' in request.POST:
                export_calls = queryset.select_related(
                    'med_card__city',
                    'med_card__district',
                    'operator'
                ).prefetch_related(
                    Prefetch(
                        'call_questions',
                        queryset=CallQuestion.objects.select_related('department').prefetch_related('questions')
                    )
                ).order_by('-created_at')
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Отчет по оператору"
                headers = [
                    '№', 'Дата и время', 'ФИО Пациента', 'Дата рождения',
                    'Адрес', 'Вопросы', 'Комментарии', 'Путь к записи', 'Оператор'
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for col_num, header_title in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header_title
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
                ws.freeze_panes = 'A2'
                for index, call in enumerate(export_calls, start=1):
                    questions_text = ""
                    if call.call_questions.exists():
                        dept_questions = {}
                        for cq in call.call_questions.all():
                            dept_name = cq.department.name if cq.department else 'Без отделения'
                            if dept_name not in dept_questions:
                                dept_questions[dept_name] = []
                            for q in cq.questions.all():
                                dept_questions[dept_name].append(q.question)
                        parts = []
                        for dept, qs in dept_questions.items():
                            if qs:
                                joined_qs = '\n- '.join(qs)
                                parts.append(f"{dept}:\n- {joined_qs}")
                        questions_text = "\n\n".join(parts)
                    row_data = [
                        index,
                        call.created_at.strftime("%d.%m.%Y %H:%M") if call.created_at else '',
                        f"{call.med_card.last_name} {call.med_card.first_name} {call.med_card.surname}" if call.med_card else '-',
                        call.med_card.birth_date.strftime("%d.%m.%Y") if call.med_card and call.med_card.birth_date else '-',
                        f"{call.med_card.city.name}, {call.med_card.district.name}" if call.med_card and call.med_card.city and call.med_card.district else (call.med_card.city.name if call.med_card and call.med_card.city else (call.med_card.district.name if call.med_card and call.med_card.district else '-')),
                        questions_text,
                        call.comment or '-',
                        call.recording_path or '-',
                        call.operator.get_full_name() if call.operator and call.operator.get_full_name() else (call.operator.username if call.operator else '-'),
                    ]
                    ws.append(row_data)
                questions_col_index = headers.index('Вопросы') + 1
                for row in ws.iter_rows(min_row=2):
                    cell = row[questions_col_index - 1]
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                from django.utils.text import slugify
                operator_name_slug = slugify(operator.get_full_name() or operator.username)
                start_str = start_date.strftime("%Y%m%d") if start_date else "all"
                end_str = end_date.strftime("%Y%m%d") if end_date else "all"
                filename = f"operator_report_{operator_name_slug}_{start_str}-{end_str}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                with BytesIO() as b:
                    wb.save(b)
                    response.write(b.getvalue())
                return response
        return render(request, 'med_app/reports.html', {
            'title': title,
            'operator_form': form,
            'questions_form': QuestionsReportForm(),
            'calls': None,
            'daily_calls': None,
        })
    return redirect('reports_url')

@custom_login_required
def questions_report_view(request):
    if request.method == 'POST':
        form = QuestionsReportForm(request.POST)
        questions_data = None
        title = "Отчет по вопросам"
        today = timezone.localtime(timezone.now()).date()
        if form.is_valid():
            department = form.cleaned_data['department']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            queryset = CallQuestion.objects.select_related(
                'department', 'call'
            ).prefetch_related(
                'questions'
            )
            if department:
                queryset = queryset.filter(department=department)
            if not start_date and not end_date:
                queryset = queryset.filter(call__created_at__date=today)
            elif start_date and not end_date:
                end_date = today
                queryset = queryset.filter(call__created_at__date__range=[start_date, end_date])
            elif start_date and end_date:
                queryset = queryset.filter(call__created_at__date__range=[start_date, end_date])
            question_stats = {}
            for cq in queryset:
                call_date = timezone.localtime(cq.call.created_at).date()
                for question in cq.questions.all():
                    key = (question.id, question.question, call_date)
                    question_stats[key] = question_stats.get(key, 0) + 1
            questions_data = [
                {
                    'id': q_id,
                    'question': q_text,
                    'date': q_date,
                    'count': count,
                    'department': department.name if department else 'Без отделения'
                }
                for (q_id, q_text, q_date), count in question_stats.items()
            ]
            questions_data.sort(key=lambda x: (x['date'], x['question']), reverse=True)
            if 'generate_report' in request.POST:
                context = {
                    'title': title,
                    'operator_form': OperatorReportForm(),
                    'questions_form': form,
                    'questions_data': questions_data,
                    'start_date': start_date,
                    'end_date': end_date,
                    'selected_department': department,
                    'today': today,
                    'visits_form': VisitsReportForm(),
                }
                return render(request, 'med_app/reports.html', context)
            elif 'export_excel' in request.POST:
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчет по вопросам"
                headers = ['№', 'Вопрос', 'Дата', 'Количество', 'Отделение']
                ws.append(headers)
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
                for index, item in enumerate(questions_data, start=1):
                    row_data = [
                        index,
                        item['question'],
                        item['date'].strftime('%d.%m.%Y'),
                        item['count'],
                        item['department']
                    ]
                    ws.append(row_data)
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"questions_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                wb.save(response)
                return response
    return render(request, 'med_app/reports.html', {
        'title': 'Отчеты',
        'operator_form': OperatorReportForm(),
        'questions_form': QuestionsReportForm(),
        'questions_data': None
    })

@custom_login_required
def visits_report_view(request):
    if request.method == 'POST':
        form = VisitsReportForm(request.POST)
        visits_data = None
        title = "Отчет по посещениям"
        today = timezone.localtime(timezone.now()).date()
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            queryset = Visit.objects.all()
            if not start_date and not end_date:
                end_date = today
                start_date = today - timedelta(days=30)
                queryset = queryset.filter(visit_time__date__range=[start_date, end_date])
            elif start_date and not end_date:
                end_date = today
                queryset = queryset.filter(visit_time__date__range=[start_date, end_date])
            elif start_date and end_date:
                queryset = queryset.filter(visit_time__date__range=[start_date, end_date])
            visits_data = queryset.values('visit_time__date').annotate(
                count=Count('id'),
                date=F('visit_time__date')
            ).order_by('date')
            if 'generate_report' in request.POST:
                context = {
                    'title': title,
                    'visits_form': form,
                    'visits_data': visits_data,
                    'start_date': start_date,
                    'end_date': end_date,
                    'today': today,
                    'operator_form': OperatorReportForm(),
                    'questions_form': QuestionsReportForm(),
                }
                return render(request, 'med_app/reports.html', context)
            elif 'export_excel' in request.POST:
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчет по посещениям"
                headers = ['№', 'Дата', 'Количество посещений']
                ws.append(headers)
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
                for index, item in enumerate(visits_data, start=1):
                    row_data = [
                        index,
                        item['date'].strftime('%d.%m.%Y'),
                        item['count']
                    ]
                    ws.append(row_data)
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"visits_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                wb.save(response)
                return response
    return redirect('reports_url')

@custom_login_required
def analytics_view(request):
    today = timezone.now().date()
    visits_hourly = Visit.objects.filter(
        visit_time__date=today
    ).annotate(
        hour=ExtractHour('visit_time')
    ).values('hour').annotate(
        count=Count('id')
    ).order_by('hour')
    visits_hourly_data = [0] * 24
    for item in visits_hourly:
        visits_hourly_data[item['hour']] = item['count']
    visits_day_parts = Visit.objects.filter(
        visit_time__date=today
    ).annotate(
        part=Case(
            When(visit_time__hour__lt=8, then=0),
            When(visit_time__hour__lt=16, then=1),
            default=2,
            output_field=IntegerField()
        )
    ).values('part').annotate(
        count=Count('id')
    ).order_by('part')
    visits_parts_data = [0, 0, 0]
    for item in visits_day_parts:
        visits_parts_data[item['part']] = item['count']
    calls_daily = Call.objects.filter(
        created_at__gte=today - timedelta(days=7)
    ).annotate(
        day=TruncDay('created_at')
    ).values('day').annotate(
        count=Count('id')
    ).order_by('day')
    calls_days = [item['day'].strftime('%d.%m') for item in calls_daily]
    calls_counts = [item['count'] for item in calls_daily]
    calls_by_operator = Call.objects.filter(
        created_at__gte=today - timedelta(days=7)
    ).values(
        'operator__username'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    context = {
        'today': today.strftime('%d.%m.%Y'),
        'visits_hours': list(range(24)),
        'visits_hourly_data': visits_hourly_data,
        'visits_parts_labels': ['Ночь (0-8)', 'День (8-16)', 'Вечер (16-24)'],
        'visits_parts_data': visits_parts_data,
        'calls_days': calls_days,
        'calls_counts': calls_counts,
        'calls_by_operator': list(calls_by_operator),
    }
    return render(request, 'med_app/analytics.html', context)

@custom_login_required
def download_wav_file(request, wav_path):
    base_dir = "/mnt/cdr"
    wav_path = wav_path.strip('/')
    if wav_path.startswith('/mnt/cdr'):
        wav_path = wav_path[8:]
    wav_path = os.path.normpath(wav_path).replace('\\', '/')
    full_path = os.path.join(base_dir, wav_path).replace('\\', '/')
    normalized_full_path = os.path.normpath(full_path)
    normalized_base_dir = os.path.normpath(base_dir)
    if not normalized_full_path.startswith(normalized_base_dir):
        return HttpResponseForbidden("Нет доступа к файлу")
    try:
        if not os.path.exists(full_path):
            return HttpResponseServerError("Файл не найден")
        if not os.access(full_path, os.R_OK):
            return HttpResponseForbidden("Нет прав на чтение файла")
        response = FileResponse(open(full_path, 'rb'), content_type='audio/wav')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
        return response
    except Exception as e:
        return HttpResponseServerError(f"Ошибка скачивания файла: {str(e)}")